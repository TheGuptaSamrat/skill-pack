#!/usr/bin/env python3
"""
HANA Volume Tracker — Excel Workbook Generator
===============================================
Reads accumulated CSV snapshots and produces an Excel workbook with:

  Sheet 1 – Raw Data          : all snapshot rows verbatim
  Sheet 2 – Table Trend       : pivot of record count & memory per table per date
  Sheet 3 – DB Size Trend     : disk usage per volume type over time
  Sheet 4 – Growth Rates      : week-over-week absolute and percentage change
  Sheet 5 – Projection        : 3- and 6-month linear extrapolation per table

Usage
-----
    pip install openpyxl
    python generate_projection_workbook.py \\
        --tables  volume-snapshots.csv \\
        --dbsize  db-size-snapshots.csv \\
        --output  projections.xlsx \\
        --horizon 26            # weeks to project forward (default 26 = 6 months)

Input CSV format
----------------
  volume-snapshots.csv  : see volume-snapshots.csv template
  db-size-snapshots.csv : see db-size-snapshots.csv template
"""

import argparse
import csv
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency — run:  pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")   # light blue
WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")   # amber
DANGER_FILL  = PatternFill("solid", fgColor="FFC7CE")   # red-ish
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=13)

THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header(ws, values: list[str], row: int = 1, col_start: int = 1) -> None:
    for c, val in enumerate(values, col_start):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _body(ws, values: list, row: int, col_start: int = 1, alt: bool = False) -> None:
    fill = ALT_FILL if alt else None
    for c, val in enumerate(values, col_start):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
        if fill:
            cell.fill = fill


def _autofit(ws, min_width: int = 10, max_width: int = 45) -> None:
    for col in ws.columns:
        width = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            min_width, min(width + 4, max_width)
        )


def _linear_regression(xs: list[float], ys: list[float]):
    """Returns (slope, intercept) for a simple OLS line through the points."""
    n = len(xs)
    if n < 2:
        return 0.0, ys[0] if ys else 0.0
    sx  = sum(xs)
    sy  = sum(ys)
    sxy = sum(x * y for x, y in zip(xs, ys))
    sxx = sum(x * x for x in xs)
    denom = n * sxx - sx * sx
    if denom == 0:
        return 0.0, sy / n
    slope     = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    return slope, intercept


def _safe_float(val: str, default: float = 0.0) -> float:
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val: str, default: int = 0) -> int:
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# Readers
# ─────────────────────────────────────────────────────────────────────────────

def read_table_snapshots(path: Path) -> list[dict]:
    rows = []
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            rows.append({
                "date":      r["SNAPSHOT_DATE"].strip(),
                "schema":    r["SCHEMA_NAME"].strip(),
                "table":     r["TABLE_NAME"].strip(),
                "records":   _safe_int(r["RECORD_COUNT"]),
                "mem_mb":    _safe_float(r["MEMORY_SIZE_MB"]),
                "disk_mb":   _safe_float(r["DISK_SIZE_MB"]),
                "delta_mb":  _safe_float(r["DELTA_MEMORY_MB"]),
                "total_mb":  _safe_float(r["TOTAL_MEMORY_MB"]),
            })
    rows.sort(key=lambda r: (r["schema"], r["table"], r["date"]))
    return rows


def read_db_snapshots(path: Path) -> list[dict]:
    rows = []
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            rows.append({
                "date":     r["SNAPSHOT_DATE"].strip(),
                "host":     r["HOST"].strip(),
                "vol_type": r["VOLUME_TYPE"].strip(),
                "used_gb":  _safe_float(r["USED_GB"]),
                "total_gb": _safe_float(r["TOTAL_GB"]),
                "util_pct": _safe_float(r["UTILIZATION_PCT"]),
            })
    rows.sort(key=lambda r: (r["vol_type"], r["date"]))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def build_raw_data(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Raw Data")
    ws.freeze_panes = "A2"
    cols = ["SNAPSHOT_DATE","SCHEMA_NAME","TABLE_NAME","RECORD_COUNT",
            "MEMORY_SIZE_MB","DISK_SIZE_MB","DELTA_MEMORY_MB","TOTAL_MEMORY_MB"]
    _header(ws, cols)
    for i, r in enumerate(rows, 2):
        _body(ws, [
            r["date"], r["schema"], r["table"],
            r["records"], r["mem_mb"], r["disk_mb"], r["delta_mb"], r["total_mb"]
        ], i, alt=bool(i % 2))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    _autofit(ws)


def build_table_trend(wb, rows: list[dict]) -> None:
    """Pivot: rows = schema.table, columns = snapshot dates (record count + memory MB)."""
    ws = wb.create_sheet("Table Trend")
    ws.freeze_panes = "C2"

    dates = sorted({r["date"] for r in rows})
    # keyed by (schema, table) -> {date: (records, total_mb)}
    pivot: dict[tuple, dict] = defaultdict(dict)
    for r in rows:
        pivot[(r["schema"], r["table"])][r["date"]] = (r["records"], r["total_mb"])

    # Header row — two sub-columns per date: Records | Memory MB
    header = ["Schema", "Table"]
    for d in dates:
        header += [f"{d} Records", f"{d} Mem MB"]
    _header(ws, header)

    for i, (key, date_map) in enumerate(sorted(pivot.items()), 2):
        row_vals = [key[0], key[1]]
        for d in dates:
            recs, mem = date_map.get(d, (0, 0.0))
            row_vals += [recs, mem]
        _body(ws, row_vals, i, alt=bool(i % 2))

    _autofit(ws)


def build_db_size_trend(wb, db_rows: list[dict]) -> None:
    ws = wb.create_sheet("DB Size Trend")
    ws.freeze_panes = "A2"
    _header(ws, ["SNAPSHOT_DATE","HOST","VOLUME_TYPE","USED_GB","TOTAL_GB","UTILIZATION_%"])
    for i, r in enumerate(db_rows, 2):
        _body(ws, [r["date"], r["host"], r["vol_type"],
                   r["used_gb"], r["total_gb"], r["util_pct"]], i, alt=bool(i % 2))
        # colour utilization
        cell = ws.cell(row=i, column=6)
        if r["util_pct"] >= 80:
            cell.fill = DANGER_FILL
        elif r["util_pct"] >= 60:
            cell.fill = WARN_FILL
    ws.auto_filter.ref = f"A1:F{len(db_rows)+1}"
    _autofit(ws)


def build_growth_rates(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Growth Rates")
    ws.freeze_panes = "A2"
    _header(ws, [
        "Schema","Table","From Date","To Date",
        "Δ Records","Δ Records %","Δ Memory MB","Δ Memory %","Days"
    ])

    # Group by (schema, table), sort by date, compute sequential deltas
    groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        groups[(r["schema"], r["table"])].append(r)

    out_rows = []
    for key, pts in sorted(groups.items()):
        pts.sort(key=lambda x: x["date"])
        for a, b in zip(pts, pts[1:]):
            try:
                d1 = datetime.strptime(a["date"], "%Y-%m-%d")
                d2 = datetime.strptime(b["date"], "%Y-%m-%d")
                days = (d2 - d1).days
            except ValueError:
                days = 0
            d_rec  = b["records"] - a["records"]
            d_mem  = round(b["total_mb"] - a["total_mb"], 2)
            pct_r  = round(d_rec  * 100 / a["records"]   if a["records"]  else 0.0, 2)
            pct_m  = round(d_mem  * 100 / a["total_mb"]  if a["total_mb"] else 0.0, 2)
            out_rows.append((key[0], key[1], a["date"], b["date"],
                             d_rec, pct_r, d_mem, pct_m, days))

    for i, r in enumerate(out_rows, 2):
        _body(ws, list(r), i, alt=bool(i % 2))
        # highlight fast growers
        pct_cell = ws.cell(row=i, column=6)
        if isinstance(r[5], float) and r[5] >= 10:
            pct_cell.fill = DANGER_FILL
        elif isinstance(r[5], float) and r[5] >= 5:
            pct_cell.fill = WARN_FILL

    ws.auto_filter.ref = f"A1:I{len(out_rows)+1}"
    _autofit(ws)


def build_projection(wb, rows: list[dict], horizon_weeks: int) -> None:
    ws = wb.create_sheet("Projection")
    ws.freeze_panes = "A2"

    groups: dict[tuple, list] = defaultdict(list)
    for r in rows:
        groups[(r["schema"], r["table"])].append(r)

    # Reference epoch for numeric regression (days since first snapshot)
    all_dates = sorted({datetime.strptime(r["date"], "%Y-%m-%d") for r in rows})
    epoch = all_dates[0]

    proj_dates = [all_dates[-1] + timedelta(weeks=w) for w in range(1, horizon_weeks + 1)]

    header = ["Schema", "Table", "Last Snapshot Date",
               "Last Records", "Last Memory MB",
               "Slope Records/day", "Slope Mem MB/day"]
    for pd in proj_dates:
        header.append(pd.strftime("%Y-%m-%d") + " Records")
    for pd in proj_dates:
        header.append(pd.strftime("%Y-%m-%d") + " Mem MB")
    _header(ws, header)

    for i, (key, pts) in enumerate(sorted(groups.items()), 2):
        pts.sort(key=lambda x: x["date"])
        xs = [(datetime.strptime(p["date"], "%Y-%m-%d") - epoch).days for p in pts]
        rec_slope, rec_int = _linear_regression(xs, [p["records"] for p in pts])
        mem_slope, mem_int = _linear_regression(xs, [p["total_mb"] for p in pts])

        last = pts[-1]
        row_vals = [
            key[0], key[1], last["date"], last["records"], last["total_mb"],
            round(rec_slope, 2), round(mem_slope, 4)
        ]
        last_x = (datetime.strptime(last["date"], "%Y-%m-%d") - epoch).days
        for pd in proj_dates:
            dx = (pd - epoch).days
            row_vals.append(max(0, round(rec_int + rec_slope * dx)))
        for pd in proj_dates:
            dx = (pd - epoch).days
            row_vals.append(max(0.0, round(mem_int + mem_slope * dx, 2)))

        _body(ws, row_vals, i, alt=bool(i % 2))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{i}"
    _autofit(ws)

    # note at bottom
    note_row = i + 2
    ws.cell(row=note_row, column=1,
            value="NOTE: These projections use a simple ordinary least-squares linear trend. "
                  "They are estimates only. Refresh monthly by re-running the SQL and regenerating.")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="595959")


def build_chart_sheet(wb, rows: list[dict]) -> None:
    """Adds a standalone chart sheet — DB DATA usage over time."""
    ws = wb.create_sheet("Chart — DB Data")

    # Build a mini table for the chart data
    data_rows = [r for r in rows if r["vol_type"] == "DATA"]
    dates   = sorted({r["date"] for r in data_rows})
    hosts   = sorted({r["host"] for r in data_rows})

    ws["A1"] = "Date"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    for hi, h in enumerate(hosts, 2):
        c = ws.cell(row=1, column=hi, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    host_map: dict[str, dict[str, float]] = {h: {} for h in hosts}
    for r in data_rows:
        host_map[r["host"]][r["date"]] = r["used_gb"]

    for ri, d in enumerate(dates, 2):
        ws.cell(row=ri, column=1, value=d)
        for hi, h in enumerate(hosts, 2):
            ws.cell(row=ri, column=hi, value=host_map[h].get(d, None))

    if len(dates) < 2:
        ws["A20"] = "Need at least 2 snapshots to draw a chart."
        return

    chart = LineChart()
    chart.title     = "HANA DATA Volume — Used GB over Time"
    chart.style     = 10
    chart.y_axis.title = "Used GB"
    chart.x_axis.title = "Snapshot Date"
    chart.height    = 14
    chart.width     = 24

    data_ref = Reference(ws, min_col=2, max_col=1 + len(hosts),
                         min_row=1, max_row=1 + len(dates))
    chart.add_data(data_ref, titles_from_data=True)

    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(dates))
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "A" + str(len(dates) + 5))
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Cover sheet
# ─────────────────────────────────────────────────────────────────────────────

def build_cover(wb, table_path: str, db_path: str, horizon_weeks: int) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50

    ws.merge_cells("B2:C2")
    title = ws["B2"]
    title.value = "HANA Volume Tracker — Projection Workbook"
    title.font  = Font(name="Calibri", bold=True, size=16, color="1F4E79")

    meta = [
        ("Generated",           datetime.today().strftime("%Y-%m-%d")),
        ("Table snapshot input", table_path),
        ("DB size input",        db_path),
        ("Projection horizon",   f"{horizon_weeks} weeks (~{horizon_weeks//4} months)"),
        ("",""),
        ("Sheet index",""),
        ("Raw Data",           "All snapshot rows from the CSV"),
        ("Table Trend",        "Pivot of record count and memory per table per date"),
        ("DB Size Trend",      "Disk usage per volume type with utilization highlights"),
        ("Growth Rates",       "Period-over-period delta and percentage change"),
        ("Projection",         "Linear extrapolation for the configured horizon"),
        ("Chart — DB Data",    "Line chart of DATA volume used over time"),
        ("",""),
        ("Refresh steps",""),
        ("1.", "Run scripts/projections/hana_volume_snapshot.sql in HANA Studio"),
        ("2.", "Append Section 1 results to volume-snapshots.csv"),
        ("3.", "Append Section 3 results to db-size-snapshots.csv"),
        ("4.", "Re-run:  python generate_projection_workbook.py"),
    ]

    for ri, (label, value) in enumerate(meta, 4):
        wb_label = ws.cell(row=ri, column=2, value=label)
        wb_value = ws.cell(row=ri, column=3, value=value)
        if label in ("Sheet index", "Refresh steps"):
            wb_label.font = Font(bold=True, size=11, color="1F4E79")
        else:
            wb_label.font = Font(size=10)
        wb_value.font = Font(size=10)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate HANA volume projection workbook from CSV snapshots."
    )
    parser.add_argument("--tables",  required=True,
                        help="Path to volume-snapshots.csv")
    parser.add_argument("--dbsize",  required=True,
                        help="Path to db-size-snapshots.csv")
    parser.add_argument("--output",  default="projections.xlsx",
                        help="Output Excel file (default: projections.xlsx)")
    parser.add_argument("--horizon", type=int, default=26,
                        help="Projection horizon in weeks (default: 26 = 6 months)")
    args = parser.parse_args()

    table_path = Path(args.tables)
    db_path    = Path(args.dbsize)
    for p in (table_path, db_path):
        if not p.exists():
            sys.exit(f"File not found: {p}")

    print("Reading snapshots …")
    table_rows = read_table_snapshots(table_path)
    db_rows    = read_db_snapshots(db_path)
    print(f"  {len(table_rows)} table rows across "
          f"{len({r['date'] for r in table_rows})} snapshots")
    print(f"  {len(db_rows)} DB-size rows across "
          f"{len({r['date'] for r in db_rows})} snapshots")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    print("Building sheets …")
    build_cover(wb, args.tables, args.dbsize, args.horizon)
    build_raw_data(wb, table_rows)
    build_table_trend(wb, table_rows)
    build_db_size_trend(wb, db_rows)
    build_growth_rates(wb, table_rows)
    build_projection(wb, table_rows, args.horizon)
    build_chart_sheet(wb, db_rows)

    out = Path(args.output)
    wb.save(out)
    print(f"\nWorkbook written → {out.resolve()}")
    print("Sheets: Cover, Raw Data, Table Trend, DB Size Trend, "
          "Growth Rates, Projection, Chart — DB Data")


if __name__ == "__main__":
    main()
